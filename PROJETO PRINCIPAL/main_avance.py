import pandas as pd
import openpyxl
from simplificador_nome import simplificador_nome
from pathlib import Path

# Faz a pesquisa no sistema dos dados de cada pessoa, trata eles e salva na planilha para ser realiza a impressão.

# It searches the system for each person's data, processes it and saves it in the spreadsheet for printing.

caminho_planilha_crachas = None


caminho_planilha_dados_colaboradores = Path(r"caminho arquivo")



data = pd.read_excel(caminho_planilha_dados_colaboradores, sheet_name="Avance", dtype={"cpf": str})
pd.set_option('display.max_colwidth', None)

posto = input("Digite o posto a seguir: ").upper()

while True:
    matricula_bruta = input("Digite a matricula a ser pegue os dados: ")
    if matricula_bruta.lower() == "s":
        break
    
    matricula = matricula_bruta.lstrip("0")

    pessoa = data.query(f"chapa == {matricula}")
    nome_bruto = pessoa["nome"].to_string(index=False)
    funcao = pessoa["funcao"].to_string(index=False)
    dt_admissao = pessoa["data_admissao"].to_string(index=False)
    chapa = pessoa["chapa"].to_string(index=False)
    rg = pessoa["identidade"].to_string(index=False)
    cpf_bruto = pessoa["cpf"].to_string(index=False)


    cpf = "{}.{}.{}-{}".format(cpf_bruto[:3], cpf_bruto[3:6], cpf_bruto[6:9], cpf_bruto[9:])
    nome = simplificador_nome(nome_bruto)
    
    


   
    caminho_planilha_crachas = Path(r"caminho arquivo")
    planilha = openpyxl.load_workbook(caminho_planilha_crachas)
    sheet = planilha["Planilha1"]

    proxima_linha = sheet.max_row + 1

    sheet.cell(row=proxima_linha, column=1, value=nome)
    sheet.cell(row=proxima_linha, column=2, value=funcao)
    sheet.cell(row=proxima_linha, column=3, value=rg)
    sheet.cell(row=proxima_linha, column=4, value=cpf)
    sheet.cell(row=proxima_linha, column=5, value=chapa)
    sheet.cell(row=proxima_linha, column=6, value=dt_admissao)
    sheet.cell(row=proxima_linha, column=7, value=posto)
    

    

    caminho_planilha_crachas = Path(r"Caminho arquivo")
    planilha.save(caminho_planilha_crachas)


print("Dados salvos com sucesso!")