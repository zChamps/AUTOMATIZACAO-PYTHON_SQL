import pandas as pd
import openpyxl
from simplificador_nome import simplificador_nome
from simplificador_escalas import escalas
from pathlib import Path
from datetime import datetime

dia_atual = f"{datetime.today().day}/{datetime.today().month}/{datetime.today().year}"



caminho_planilha_dados_colaboradores = Path(r"caminho do arquivo com os dados dos colaboradores")


data = pd.read_excel(caminho_planilha_dados_colaboradores, dtype={"pispasep": str})
pd.set_option('display.max_colwidth', None)
while True:
    matricula_bruta = input("Digite a matricula a ser pegue os dados: ")
    if matricula_bruta.lower() == "s":
        break
    
    matricula = matricula_bruta.lstrip("0")

    pessoa = data.query(f"chapa == {matricula}")
    chapa = pessoa["chapa"].to_string(index=False)
    nome_bruto = pessoa["nome"].to_string(index=False)
    dt_admissao = pessoa["data_admissao"].to_string(index=False)
    funcao = pessoa["funcao"].to_string(index=False)
    secao = pessoa["secao"].to_string(index=False)
    escala_bruta = pessoa["escala"].to_string(index=False)
    ctps_bruto = pessoa["ctps"].to_string(index=False)
    pis = pessoa["pispasep"].to_string(index=False)
    rg = pessoa["identidade"].to_string(index=False)


    nome = simplificador_nome(nome_bruto)
    escala = escalas(escala_bruta)
    ctps = ctps_bruto.split("-")[0]



    caminho_planilha_crachas = Path(r"caminho da planilha onde serão colocados os dados para realizar a impressão dos crachás")

    planilha = openpyxl.load_workbook(caminho_planilha_crachas)
    sheet = planilha["Plan1"]

    proxima_linha = sheet.max_row + 1

    sheet.cell(row=proxima_linha, column=1, value=chapa)
    sheet.cell(row=proxima_linha, column=2, value=nome)
    sheet.cell(row=proxima_linha, column=3, value=rg)
    sheet.cell(row=proxima_linha, column=4, value=ctps)
    sheet.cell(row=proxima_linha, column=5, value=pis)
    sheet.cell(row=proxima_linha, column=6, value=escala)
    sheet.cell(row=proxima_linha, column=7, value=dt_admissao)
    sheet.cell(row=proxima_linha, column=8, value=funcao)
    sheet.cell(row=proxima_linha, column=9, value=secao)
    sheet.cell(row=proxima_linha, column=10, value="SIM")
    
    sheet.cell(row=proxima_linha, column=12, value=dia_atual)
    sheet.cell(row=proxima_linha, column=13, value="WILLIAN")

    # Salvar o arquivo
    planilha.save(caminho_planilha_crachas)

print("Dados salvos com sucesso!")
