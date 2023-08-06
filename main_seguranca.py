import pandas as pd
import openpyxl
from simplificador_nome import simplificador_nome
from pathlib import Path
from datetime import datetime

dia_atual = f"{datetime.today().day}/{datetime.today().month}/{datetime.today().year}"
caminho_planilha_crachas = None
empresa = input("Digite a empresa onde será salvo os dados dos colaboradores: ").upper()

caminho_planilha_dados_colaboradores = Path(r"caminho do arquivo com os dados dos colaboradores")
if empresa == "SERVIARM":
        nome_planilha = "Serviarm"
elif empresa == "SERVAL SEGURANÇA" or empresa == "SEGURANÇA":
        nome_planilha = "Serval Segurança"


data = pd.read_excel(caminho_planilha_dados_colaboradores, sheet_name=nome_planilha)
pd.set_option('display.max_colwidth', None)


while True:
    matricula_bruta = input("Digite a matricula a ser pegue os dados: ")
    if matricula_bruta.lower() == "s":
        break
    
    matricula = matricula_bruta.lstrip("0")

    pessoa = data.query(f"chapa == {matricula}")
    fator_rh = pessoa["tipo_sanguineo"].to_string(index=False)
    nome_bruto = pessoa["nome"].to_string(index=False)
    funcao = pessoa["funcao"].to_string(index=False)
    dt_admissao = pessoa["data_admissao"].to_string(index=False)
    chapa = pessoa["chapa"].to_string(index=False)
    registro_pf = pessoa["registro_pf"].to_string(index=False)
    cpf = pessoa["cpf"].to_string(index=False)
    reciclagem = pessoa["data_reciclagem"].to_string(index=False)


    nome = simplificador_nome(nome_bruto)
    
    


    if empresa == "SERVIARM":
        caminho_planilha_crachas = Path(r"caminho da planilha onde serão colocados os dados para realizar a impressão dos crachás")
    elif empresa == "SERVAL SEGURANÇA" or empresa == "SEGURANÇA":
        caminho_planilha_crachas = Path(r"caminho da planilha onde serão colocados os dados para realizar a impressão dos crachás")
    planilha = openpyxl.load_workbook(caminho_planilha_crachas)
    sheet = planilha["Planilha1"]

    proxima_linha = sheet.max_row + 1

    sheet.cell(row=proxima_linha, column=1, value=fator_rh)
    sheet.cell(row=proxima_linha, column=2, value=nome)
    sheet.cell(row=proxima_linha, column=3, value=funcao)
    sheet.cell(row=proxima_linha, column=4, value=dt_admissao)
    sheet.cell(row=proxima_linha, column=5, value=chapa)
    sheet.cell(row=proxima_linha, column=6, value=dia_atual)
    sheet.cell(row=proxima_linha, column=8, value=registro_pf)
    sheet.cell(row=proxima_linha, column=9, value=cpf)
    sheet.cell(row=proxima_linha, column=10, value=reciclagem)

    # Salvar o arquivo
    
    if empresa == "SERVIARM":
        caminho_planilha_crachas = Path(r"caminho da planilha onde serão colocados os dados para realizar a impressão dos crachás")
        planilha.save(caminho_planilha_crachas)
    elif empresa == "SERVAL SEGURANÇA" or empresa == "SEGURANÇA":
        caminho_planilha_crachas = Path(r"caminho da planilha onde serão colocados os dados para realizar a impressão dos crachás")
        planilha.save(caminho_planilha_crachas)

print("Dados salvos com sucesso!")